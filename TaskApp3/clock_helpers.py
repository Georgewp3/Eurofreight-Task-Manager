from datetime import date, datetime, time, timedelta
import re

from models import (
    db,
    User,
    ContractModel,
    ContractModelDay,
    UserContractAssignment,
    ClockRecord,
    ClockExtraInstruction,
)


WEEKDAY_LABELS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
DEFAULT_CLOCK_START = date(2026, 5, 20)
DEFAULT_CLOCK_END = date(2026, 6, 18)

# Current demo workbook layout. Keep this small mapping easy to change later.
CLOCK_IMPORT_COLUMNS = {
    "name": 0,
    "date": 1,
    "weekday": 2,
    "clock_in": 3,
    "clock_out": 4,
}

NAME_ALIASES = {
    "valentinos melinioti": "Valentinos Meliniotis",
    "merlin basseck": "Merlin Basseck Noah",
    "christopher mbu": "Mbu Christopher Bate",
    "lih": "Lih Rostent",
    "tzeni ntima": "Tzeni Dima",
    "elena giallourou": "Elena Yiallourou",
    "ravinderjit singh": "Ravinderjit Singh (Jot)",
    "gourav": "GOURAV",
    "rajinder kumar sabi": "RAJINDER KUMAR SABI",
    "sabi rajinder kumar": "RAJINDER KUMAR SABI",
    "lovepreet singh": "LOVEPREET SINGH",
    "sandeep singh": "Sandeep Singh (in cargo with Picka)",
    "narender singh": "Narender Singh (in pop with Marina)",
    "navin koumar": "NAVIN KUMAR",
    "samsher singh": "SAMSHER SINGH",
    "samher singh": "SAMSHER SINGH",
    "ramandeep singh": "Ramandeep Singh (Raman)",
    "pailak tartarian": "Pailak Tatarian",
}

CONTRACT_SEEDS = {
    "Contract Model 1": {
        "users": ["Valentinos Meliniotis", "Elena Toumazou", "Pailak Tatarian"],
        "days": [
            ("06:00", "18:00", 90, 630, False),
            ("06:30", "17:00", 90, 540, False),
            ("06:30", "16:30", 90, 510, False),
            ("06:30", "16:30", 90, 510, False),
            ("06:30", "16:30", 90, 510, False),
            ("06:30", "13:00", 90, 300, False),
            (None, None, 0, 0, True),
        ],
    },
    "Contract Model 2": {
        "users": ["Merlin Basseck Noah", "Bodylawson", "Mbu Christopher Bate", "Lih Rostent"],
        "days": [
            ("06:00", "20:00", 90, 750, False),
            ("06:00", "17:00", 90, 570, False),
            ("06:00", "16:30", 90, 540, False),
            ("06:00", "16:30", 90, 540, False),
            ("06:00", "16:30", 90, 540, False),
            ("06:30", "14:00", 90, 360, False),
            (None, None, 0, 0, True),
        ],
    },
    "Contract Model 3": {
        "users": ["Tzeni Dima", "Elena Yiallourou", "Marina Aspromalli", "Anna Maria"],
        "days": [
            ("08:00", "20:00", 90, 630, False),
            ("08:00", "17:30", 90, 480, False),
            ("08:00", "17:30", 90, 480, False),
            ("08:00", "17:30", 90, 480, False),
            ("08:00", "17:30", 90, 480, False),
            ("08:00", "14:00", 90, 270, False),
            (None, None, 0, 0, True),
        ],
    },
    "Contract Model 4": {
        "users": [
            "Ravinderjit Singh (Jot)",
            "GOURAV",
            "RAJINDER KUMAR SABI",
            "LOVEPREET SINGH",
            "Sandeep Singh (in cargo with Picka)",
            "Narender Singh (in pop with Marina)",
            "NAVIN KUMAR",
            "SAMSHER SINGH",
            "Ramandeep Singh (Raman)",
        ],
        "days": [
            ("06:30", "17:30", 90, 570, False),
            ("06:30", "17:30", 90, 570, False),
            ("06:30", "17:30", 90, 570, False),
            ("06:30", "17:30", 90, 570, False),
            ("06:30", "17:30", 90, 570, False),
            ("06:30", "14:30", 30, 450, False),
            (None, None, 0, 0, True),
        ],
    },
}


def normalize_name(value):
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def parse_hhmm(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)):
        minutes = round(float(value) * 24 * 60)
        return f"{minutes // 60:02d}:{minutes % 60:02d}"

    raw = str(value).strip()
    if not raw or raw == "-":
        return None
    match = re.search(r"(\d{1,2}):(\d{2})", raw)
    if not match:
        return None
    hours = int(match.group(1))
    minutes = int(match.group(2))
    if hours > 23 or minutes > 59:
        return None
    return f"{hours:02d}:{minutes:02d}"


def parse_work_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()

    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    return None


def minutes_between(clock_in, clock_out):
    start = parse_hhmm(clock_in)
    end = parse_hhmm(clock_out)
    if not start or not end:
        return None
    start_dt = datetime.strptime(start, "%H:%M")
    end_dt = datetime.strptime(end, "%H:%M")
    if end_dt < start_dt:
        end_dt += timedelta(days=1)
    return int((end_dt - start_dt).total_seconds() // 60)


def instruction_scope_label(instruction):
    if instruction.scope_type == "all":
        return "All users"
    if instruction.scope_type == "contract_model" and instruction.contract_model:
        return f"Contract model: {instruction.contract_model.name}"
    if instruction.scope_type in {"user", "users"}:
        ids = instruction_user_ids(instruction)
        if ids:
            users = User.query.filter(User.id.in_(ids)).order_by(User.full_name.asc()).all()
            names = ", ".join(user.full_name for user in users)
            return f"Users: {names}" if names else "Specific users"
        if instruction.user:
            return f"User: {instruction.user.full_name}"
        return "Specific users"
    return instruction.scope_type


def instruction_user_ids(instruction):
    ids = []
    if instruction.user_ids:
        for raw_id in instruction.user_ids.split(","):
            raw_id = raw_id.strip()
            if raw_id.isdigit():
                ids.append(int(raw_id))
    elif instruction.user_id:
        ids.append(instruction.user_id)
    return ids


def format_minutes(minutes):
    if minutes is None:
        return "-"
    sign = "-" if minutes < 0 else ""
    minutes = abs(int(round(minutes)))
    return f"{sign}{minutes // 60}:{minutes % 60:02d}"


def seed_contract_models():
    for model_name, payload in CONTRACT_SEEDS.items():
        model = ContractModel.query.filter_by(name=model_name).first()
        if not model:
            model = ContractModel(name=model_name, active=True)
            db.session.add(model)
            db.session.flush()

        model.active = True
        existing_days = {day.weekday: day for day in model.days}
        for weekday, day_payload in enumerate(payload["days"]):
            start_time, end_time, break_minutes, expected_minutes, is_flat_pay = day_payload
            day_row = existing_days.get(weekday)
            if not day_row:
                day_row = ContractModelDay(contract_model=model, weekday=weekday)
                db.session.add(day_row)
            day_row.start_time = start_time
            day_row.end_time = end_time
            day_row.break_minutes = break_minutes
            day_row.expected_minutes = expected_minutes
            day_row.is_flat_pay = is_flat_pay

        for full_name in payload["users"]:
            user = User.query.filter_by(full_name=full_name).first()
            if not user:
                continue
            exists = UserContractAssignment.query.filter_by(
                user_id=user.id,
                contract_model_id=model.id,
                active=True,
            ).first()
            if exists:
                continue

            for assignment in UserContractAssignment.query.filter_by(
                user_id=user.id,
                active=True,
            ).all():
                assignment.active = False

            historical = UserContractAssignment.query.filter_by(
                user_id=user.id,
                contract_model_id=model.id,
                active=False,
            ).first()
            if historical:
                historical.active = True
            else:
                db.session.add(UserContractAssignment(user=user, contract_model=model))

    db.session.commit()


def resolve_user_from_name(raw_name):
    normalized = normalize_name(raw_name)
    target_name = NAME_ALIASES.get(normalized, raw_name)
    lookup = {normalize_name(user.full_name): user for user in User.query.all()}
    return lookup.get(normalize_name(target_name))


def import_clock_workbook(file_storage):
    from openpyxl import load_workbook

    source = getattr(file_storage, "stream", file_storage)
    workbook = load_workbook(source, data_only=True, read_only=True)
    sheet = workbook.active
    imported = 0
    updated = 0
    skipped = 0
    unmatched = {}

    for row in sheet.iter_rows(values_only=True):
        if not row or all(cell is None for cell in row):
            skipped += 1
            continue

        raw_name = row[CLOCK_IMPORT_COLUMNS["name"]] if len(row) > 0 else None
        if normalize_name(raw_name) in {"name", "employee", "employee name", "user"}:
            skipped += 1
            continue

        work_date = parse_work_date(row[CLOCK_IMPORT_COLUMNS["date"]] if len(row) > 1 else None)
        user = resolve_user_from_name(raw_name)
        if not user:
            if raw_name:
                key = str(raw_name).strip()
                unmatched[key] = unmatched.get(key, 0) + 1
            skipped += 1
            continue
        if not work_date:
            skipped += 1
            continue

        clock_in = parse_hhmm(row[CLOCK_IMPORT_COLUMNS["clock_in"]] if len(row) > 3 else None)
        clock_out = parse_hhmm(row[CLOCK_IMPORT_COLUMNS["clock_out"]] if len(row) > 4 else None)
        record = ClockRecord.query.filter_by(user_id=user.id, work_date=work_date).first()
        if record:
            record.clock_in = clock_in
            record.clock_out = clock_out
            record.source = "excel_import"
            updated += 1
        else:
            db.session.add(ClockRecord(
                user=user,
                work_date=work_date,
                clock_in=clock_in,
                clock_out=clock_out,
                source="excel_import",
            ))
            imported += 1

    db.session.commit()
    return {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "unmatched": unmatched,
    }


def get_active_assignment(user_id, work_date=None):
    assignments = UserContractAssignment.query.filter_by(user_id=user_id, active=True).all()
    if work_date is None:
        return assignments[0] if assignments else None
    for assignment in assignments:
        if assignment.effective_from and assignment.effective_from > work_date:
            continue
        if assignment.effective_to and assignment.effective_to < work_date:
            continue
        return assignment
    return assignments[0] if assignments else None


def get_contract_day_for_user(user_id, work_date):
    assignment = get_active_assignment(user_id, work_date)
    if not assignment:
        return None, None
    contract_day = ContractModelDay.query.filter_by(
        contract_model_id=assignment.contract_model_id,
        weekday=work_date.weekday(),
    ).first()
    return assignment, contract_day


def instruction_applies_to_user(instruction, user_id, assignment):
    if instruction.scope_type == "all":
        return True
    if instruction.scope_type == "contract_model":
        return bool(
            assignment
            and instruction.contract_model_id
            and assignment.contract_model_id == instruction.contract_model_id
        )
    if instruction.scope_type in {"user", "users"}:
        return user_id in instruction_user_ids(instruction)
    return False


def calculate_clock_summary(user_id, start_date, end_date):
    user = User.query.get(user_id) if user_id else None
    if not user:
        return None

    records = ClockRecord.query.filter(
        ClockRecord.user_id == user.id,
        ClockRecord.work_date >= start_date,
        ClockRecord.work_date <= end_date,
    ).all()
    records_by_date = {record.work_date: record for record in records}

    rows = []
    total_expected = 0
    total_actual = 0
    total_overtime = 0
    total_less = 0
    total_difference = 0
    total_boss_extra_pay = 0.0
    cursor = start_date
    instructions = ClockExtraInstruction.query.filter(
        ClockExtraInstruction.active.is_(True),
        ClockExtraInstruction.work_date >= start_date,
        ClockExtraInstruction.work_date <= end_date,
    ).all()
    instructions_by_date = {}
    for instruction in instructions:
        instructions_by_date.setdefault(instruction.work_date, []).append(instruction)

    while cursor <= end_date:
        assignment, contract_day = get_contract_day_for_user(user.id, cursor)
        record = records_by_date.get(cursor)
        expected = contract_day.expected_minutes if contract_day else 0
        break_minutes = contract_day.break_minutes if contract_day else 0
        actual_gross = minutes_between(record.clock_in, record.clock_out) if record else None
        actual_net = 0
        overtime = 0
        less = 0
        difference = 0
        status = "No clock record"
        excluded = False
        matching_instructions = [
            instruction
            for instruction in instructions_by_date.get(cursor, [])
            if instruction_applies_to_user(instruction, user.id, assignment)
        ]
        boss_instruction_title = ", ".join(i.title for i in matching_instructions) or "-"
        boss_extra_minutes = 0
        boss_extra_pay = 0.0

        if contract_day and contract_day.is_flat_pay:
            status = "Flat pay day - excluded from demo calculation"
            excluded = True
        elif not assignment:
            status = "No active contract assignment"
        elif not contract_day:
            status = "No contract day rule"
        elif record and actual_gross is not None:
            actual_net = max(actual_gross - break_minutes, 0)
            difference = actual_net - expected
            overtime = max(difference, 0)
            less = abs(min(difference, 0))
            status = "OK"
        elif record:
            status = "Incomplete clock record"

        if (
            matching_instructions
            and not excluded
            and record
            and record.clock_in
            and record.clock_out
            and contract_day
            and contract_day.end_time
        ):
            boss_extra_minutes = max(minutes_between(contract_day.end_time, record.clock_out) or 0, 0)
            # Demo formula: approved minutes are counted once, then multiplied by each
            # matching extra rate. This keeps approved time separate from normal overtime.
            boss_extra_pay = sum(
                (boss_extra_minutes / 60) * instruction.extra_rate_per_hour
                for instruction in matching_instructions
            )
            if boss_extra_minutes == 0 and status == "OK":
                status = "Instruction applied, but no extra time after scheduled end"
        total_boss_extra_pay += boss_extra_pay

        if not excluded:
            total_expected += expected
            total_actual += actual_net
            total_overtime += overtime
            total_less += less
            total_difference += difference

        rows.append({
            "date": cursor,
            "day": WEEKDAY_LABELS[cursor.weekday()],
            "contract_model": assignment.contract_model.name if assignment else "-",
            "expected": expected if not excluded else None,
            "clock_in": record.clock_in if record else None,
            "clock_out": record.clock_out if record else None,
            "actual_net": actual_net if not excluded else None,
            "overtime": overtime if not excluded else None,
            "less": less if not excluded else None,
            "difference": difference if not excluded else None,
            "boss_instruction": boss_instruction_title,
            "boss_extra_minutes": boss_extra_minutes,
            "boss_extra_pay": boss_extra_pay,
            "status": status,
            "status_class": "status-ok" if status == "OK" else "status-warning",
        })
        cursor += timedelta(days=1)

    assignment = get_active_assignment(user.id, start_date)
    hourly_rate = assignment.hourly_rate if assignment else 0.0
    # Normal demo formula: payable positive net only. This can change when payroll
    # rules are finalized.
    owed = max(total_overtime - total_less, 0) / 60 * hourly_rate
    # Boss-approved extra pay is intentionally layered on top of normal estimated
    # pay and does not modify contract models or subtract break time again.
    grand_owed = owed + total_boss_extra_pay

    return {
        "user": user,
        "assignment": assignment,
        "record_count": len(records),
        "rows": rows,
        "summary": {
            "total_expected": total_expected,
            "total_actual": total_actual,
            "total_overtime": total_overtime,
            "total_less": total_less,
            "net_difference": total_difference,
            "hourly_rate": hourly_rate,
            "owed": owed,
            "boss_extra_pay": total_boss_extra_pay,
            "grand_owed": grand_owed,
        },
    }


def contract_model_overview():
    return ContractModel.query.order_by(ContractModel.name.asc()).all()


def recent_clock_records(limit=12):
    return (
        ClockRecord.query
        .join(User)
        .order_by(ClockRecord.work_date.desc(), User.full_name.asc())
        .limit(limit)
        .all()
    )
